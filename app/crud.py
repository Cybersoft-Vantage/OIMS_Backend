from sqlalchemy import func, text
from sqlalchemy.exc import ProgrammingError
from sqlalchemy.orm import Session
from fastapi import UploadFile
from . import models, security
from .schemas import (
    EmployeeDetailCreate,
    EmployeeDetailUpdate,
    CategoryCreate,
    CategoryUpdate,
    SubCategoryCreate,
    SubCategoryUpdate,
    AssetStatusCreate,
    AssetStatusUpdate,
    AssetCreate,
    AssetUpdate,
    AssetComponentCreate,
    AssetComponentUpdate,
    AssetAssignmentCreate,
    AssetAssignmentReturn,
    AssetHistoryCreate,
    DetailedAssetAssignmentCreate,
    DetailedAssetAssignmentReturn,
    DetailedAssetAssignmentOut,
    DetailedAssetHistoryCreate,
    DetailedCategoryCreate,
    DetailedCategoryUpdate,
    DetailedAssetCreate,
    DetailedAssetUpdate,
    ProcurementRequestCreate,
    ProcurementRequestUpdate,
    SoftwareLicenseCreate,
    SoftwareLicenseUpdate,
)
import datetime
import re

ASSIGNED_STATUS = "Assigned"
AVAILABLE_STATUS = "Available"
MAINTENANCE_STATUS = "Maintenance"
# Statuses that put an asset out of service: it cannot be issued to anyone while its
# status reads like this. Matched loosely so "Under Maintenance", "Damaged" and "Under
# Repair" are all caught, not just the exact words.
OUT_OF_SERVICE_STATUS_PATTERN = re.compile(r"damag|maintenance|repair|sold", re.IGNORECASE)


def is_asset_out_of_service(status: str | None) -> bool:
    """True when a status means the asset is not available to be issued."""
    return bool(OUT_OF_SERVICE_STATUS_PATTERN.search(str(status or "")))


def get_employee_detail(db: Session, employee_id: int):
    return db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == employee_id).first()


def get_employee_detail_by_user_id(db: Session, user_id: str):
    return db.query(models.EmployeeDetail).filter(models.EmployeeDetail.UserId == user_id).first()


def get_employee_detail_by_identifier(db: Session, identifier: str):
    normalized = (identifier or "").strip()
    if not normalized:
        return None
    return (
        db.query(models.EmployeeDetail)
        .filter((models.EmployeeDetail.UserId == normalized) | (models.EmployeeDetail.Email == normalized))
        .first()
    )


def get_employee_details(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.EmployeeDetail).offset(skip).limit(limit).all()


def create_employee_detail(db: Session, employee_in: EmployeeDetailCreate):
    data = employee_in.model_dump(exclude={"Password"})
    password = (employee_in.Password or "").strip()
    if password:
        data["PasswordHash"] = security.get_password_hash(password)
    employee = models.EmployeeDetail(**data)
    db.add(employee)
    db.commit()
    db.refresh(employee)
    return employee


def update_employee_detail(db: Session, employee_id: int, employee_in: EmployeeDetailUpdate):
    employee = get_employee_detail(db, employee_id)
    if not employee:
        return None

    original_email = employee.Email
    data = employee_in.model_dump(exclude={"Password"}, exclude_unset=True)
    for field, value in data.items():
        setattr(employee, field, value)

    email_changed = data.get("Email") is not None and data.get("Email") != original_email
    if email_changed:
        employee.Verify = 0
        employee.EmailVerifiedAt = None
        employee.VerificationOtpHash = None
        employee.VerificationOtpExpiresAt = None
        employee.VerificationOtpAttempts = 0
        employee.VerificationOtpSentAt = None

    db.commit()
    db.refresh(employee)

    password = (employee_in.Password or "").strip() if employee_in.Password is not None else ""
    if password:
        employee.PasswordHash = security.get_password_hash(password)
        db.commit()
        db.refresh(employee)

    return employee


def invalidate_employee_verification(db: Session, employee: models.EmployeeDetail):
    employee.Verify = 0
    employee.EmailVerifiedAt = None
    employee.VerificationOtpHash = None
    employee.VerificationOtpExpiresAt = None
    employee.VerificationOtpAttempts = 0
    employee.VerificationOtpSentAt = None
    db.commit()
    db.refresh(employee)
    return employee


def delete_employee_detail(db: Session, employee_id: int):
    employee = get_employee_detail(db, employee_id)
    if not employee:
        return None
    db.delete(employee)
    db.commit()
    return employee


def import_employee_details(db: Session, file: UploadFile):
    import csv
    from io import StringIO

    field_map = {
        "userid": "UserId",
        "user_id": "UserId",
        "fullname": "FullName",
        "full_name": "FullName",
        "department": "Department",
        "designation": "Designation",
        "email": "Email",
        "phone": "Phone",
        "role": "Role",
        "isactive": "IsActive",
        "is_active": "IsActive",
        "password": "Password",
    }
    valid_roles = {"employee", "admin", "hr"}

    file_extension = file.filename.rsplit(".", 1)[-1].lower() if file.filename and "." in file.filename else ""
    rows = []
    if file_extension == "csv":
        text_data = file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text_data))
        rows = list(reader)
    elif file_extension in {"xls", "xlsx"}:
        try:
            import openpyxl
        except ImportError:
            raise ValueError("XLS/XLSX import requires openpyxl. Please install it before using this feature.")

        workbook = openpyxl.load_workbook(file.file, data_only=True)
        sheet = workbook.active
        header = [str(cell).strip() if cell is not None else "" for cell in next(sheet.iter_rows(values_only=True))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({header[idx].strip(): row[idx] for idx in range(len(header)) if idx < len(row)})
    else:
        raise ValueError("Unsupported file type. Please upload a CSV or XLSX file.")

    processed = 0
    created = 0
    updated = 0
    errors: list[dict] = []

    for index, raw_row in enumerate(rows, start=1):
        processed += 1
        try:
            row = {
                key.strip().lower(): (value if value is None or isinstance(value, str) else str(value))
                for key, value in raw_row.items()
                if key
            }
            payload: dict = {}
            for raw_key, raw_value in row.items():
                target = field_map.get(raw_key)
                if not target:
                    continue
                value = str(raw_value).strip() if raw_value is not None else None
                if target == "IsActive":
                    if value in (None, ""):
                        payload[target] = 1
                    elif value.lower() in {"1", "true", "yes", "active"}:
                        payload[target] = 1
                    elif value.lower() in {"0", "false", "no", "inactive"}:
                        payload[target] = 0
                    else:
                        raise ValueError("IsActive must be one of: 1/0/true/false/yes/no/active/inactive")
                else:
                    payload[target] = value

            user_id = (payload.get("UserId") or "").strip()
            full_name = (payload.get("FullName") or "").strip()
            if not user_id:
                raise ValueError("Missing required field 'UserId'.")
            if not full_name:
                raise ValueError("Missing required field 'FullName'.")

            role = (payload.get("Role") or "employee").strip().lower()
            if role not in valid_roles:
                raise ValueError("Role must be one of: employee, admin, hr")
            payload["Role"] = role

            password = (payload.get("Password") or "").strip()
            payload["Password"] = password or "CSV112233"

            existing = get_employee_detail_by_user_id(db, user_id)
            if existing:
                update_payload = EmployeeDetailUpdate(**payload)
                updated_employee = update_employee_detail(db, existing.EmployeeId, update_payload)
                if not updated_employee:
                    raise ValueError(f"Unable to update user with UserId '{user_id}'")
                updated += 1
            else:
                create_payload = EmployeeDetailCreate(**payload)
                create_employee_detail(db, create_payload)
                created += 1
        except Exception as exc:
            db.rollback()
            errors.append({"row": index, "error": str(exc)})

    return {"processed": processed, "created": created, "updated": updated, "errors": errors}


def get_category(db: Session, category_id: int):
    return db.query(models.Category).filter(models.Category.CategoryId == category_id).first()


def get_categories(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Category).filter(models.Category.IsDeleted == 0).offset(skip).limit(limit).all()


def create_category(db: Session, category_in: CategoryCreate):
    category = models.Category(**category_in.model_dump())
    db.add(category)
    db.commit()
    db.refresh(category)
    return category


def update_category(db: Session, category_id: int, category_in: CategoryUpdate):
    category = get_category(db, category_id)
    if not category:
        return None
    for field, value in category_in.model_dump(exclude_unset=True).items():
        setattr(category, field, value)
    db.commit()
    db.refresh(category)
    return category


def delete_category(db: Session, category_id: int):
    category = get_category(db, category_id)
    if not category:
        return None
    
    # Check if category is in use
    asset_count = db.query(models.Asset).filter(models.Asset.CategoryId == category_id).count()
    if asset_count > 0:
        return {"error": f"Cannot delete category. It is assigned to {asset_count} asset(s)."}
    
    # Soft delete
    category.IsDeleted = 1
    category.DeletedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(category)
    return category


def get_subcategory(db: Session, subcategory_id: int):
    return db.query(models.SubCategory).filter(models.SubCategory.SubCategoryId == subcategory_id).first()


def get_subcategories(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.SubCategory).filter(models.SubCategory.IsDeleted == 0).offset(skip).limit(limit).all()


def create_subcategory(db: Session, subcategory_in: SubCategoryCreate):
    subcategory = models.SubCategory(**subcategory_in.model_dump())
    db.add(subcategory)
    db.commit()
    db.refresh(subcategory)
    return subcategory



def _generate_next_procurement_reference(db: Session) -> str:
    # Find highest existing numeric suffix for references like 'PR-001'
    rows = db.query(models.ProcurementRequest.Reference).all()
    max_num = 0
    for (ref,) in rows:
        if not ref:
            continue
        if ref.upper().startswith('PR-'):
            try:
                num = int(ref.split('-')[-1])
                if num > max_num:
                    max_num = num
            except Exception:
                continue
    return f"PR-{(max_num + 1):03d}"


def _generate_next_workorder_reference(db: Session) -> str:
    rows = db.query(models.WorkOrder.Reference).all()
    max_num = 0
    for (ref,) in rows:
        if not ref:
            continue
        if ref.upper().startswith('WO-'):
            try:
                num = int(ref.split('-')[-1])
                if num > max_num:
                    max_num = num
            except Exception:
                continue
    return f"WO-{(max_num + 1):04d}"


def get_procurement(db: Session, procurement_id: int):
    return db.query(models.ProcurementRequest).filter(models.ProcurementRequest.ProcurementId == procurement_id).first()


def get_procurements(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.ProcurementRequest).offset(skip).limit(limit).all()


def create_procurement(db: Session, procurement_in: ProcurementRequestCreate):
    data = procurement_in.model_dump()
    category_id = data.get('CategoryId')
    subcategory_id = data.get('SubCategoryId')

    if category_id is not None and not get_detailed_category(db, category_id):
        raise ValueError(f"Detailed category not found: {category_id}")

    if subcategory_id is not None:
        subcategory = get_detailed_category(db, subcategory_id)
        if not subcategory:
            raise ValueError(f"Detailed subcategory not found: {subcategory_id}")
        if category_id is not None and subcategory.ParentId != category_id:
            raise ValueError(
                f"Subcategory {subcategory_id} does not belong to detailed category {category_id}"
            )
        if category_id is None:
            data['CategoryId'] = subcategory.ParentId

    if not data.get('Reference'):
        data['Reference'] = _generate_next_procurement_reference(db)
    proc = models.ProcurementRequest(**data)
    db.add(proc)
    db.commit()
    db.refresh(proc)
    return proc


def update_procurement(db: Session, procurement_id: int, procurement_in: ProcurementRequestUpdate):
    proc = get_procurement(db, procurement_id)
    if not proc:
        return None

    update_data = procurement_in.model_dump(exclude_unset=True)
    category_id = update_data.get('CategoryId', proc.CategoryId)
    subcategory_id = update_data.get('SubCategoryId', proc.SubCategoryId)

    if category_id is not None and not get_detailed_category(db, category_id):
        raise ValueError(f"Detailed category not found: {category_id}")

    if subcategory_id is not None:
        subcategory = get_detailed_category(db, subcategory_id)
        if not subcategory:
            raise ValueError(f"Detailed subcategory not found: {subcategory_id}")
        if category_id is not None and subcategory.ParentId != category_id:
            raise ValueError(
                f"Subcategory {subcategory_id} does not belong to detailed category {category_id}"
            )
        if category_id is None:
            update_data['CategoryId'] = subcategory.ParentId

    for field, value in update_data.items():
        setattr(proc, field, value)
    db.commit()
    db.refresh(proc)
    return proc


def delete_procurement(db: Session, procurement_id: int):
    proc = get_procurement(db, procurement_id)
    if not proc:
        return None
    db.delete(proc)
    db.commit()
    return proc


def get_licenses(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.SoftwareLicense).offset(skip).limit(limit).all()


def create_license(db: Session, license_in: SoftwareLicenseCreate):
    license = models.SoftwareLicense(**license_in.model_dump())
    db.add(license)
    db.commit()
    db.refresh(license)
    return license


def update_license(db: Session, license_id: int, license_in: SoftwareLicenseUpdate):
    license = db.query(models.SoftwareLicense).filter(models.SoftwareLicense.LicenseId == license_id).first()
    if not license:
        return None
    for field, value in license_in.model_dump(exclude_unset=True).items():
        setattr(license, field, value)
    db.commit()
    db.refresh(license)
    return license


def delete_license(db: Session, license_id: int):
    license = db.query(models.SoftwareLicense).filter(models.SoftwareLicense.LicenseId == license_id).first()
    if not license:
        return None
    db.delete(license)
    db.commit()
    return license


def update_subcategory(db: Session, subcategory_id: int, subcategory_in: SubCategoryUpdate):


    subcategory = get_subcategory(db, subcategory_id)
    if not subcategory:


        return None
    for field, value in subcategory_in.model_dump(exclude_unset=True).items():
        setattr(subcategory, field, value)
    db.commit()
    db.refresh(subcategory)
    return subcategory


def delete_subcategory(db: Session, subcategory_id: int):
    subcategory = get_subcategory(db, subcategory_id)


    if not subcategory:
        return None
    
    # Check if subcategory is in use
    asset_count = db.query(models.Asset).filter(models.Asset.SubCategoryId == subcategory_id).count()
    component_count = db.query(models.AssetComponent).filter(models.AssetComponent.SubCategoryId == subcategory_id).count()
    
    if asset_count > 0 or component_count > 0:
        return {


            "error": f"Cannot delete subcategory. It is assigned to {asset_count} asset(s) and {component_count} component(s)."
        }
    
    # Soft delete
    subcategory.IsDeleted = 1
    subcategory.DeletedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(subcategory)
    return subcategory


def get_asset_status(db: Session, status_id: int):
    return db.query(models.AssetStatus).filter(models.AssetStatus.StatusId == status_id).first()


def get_asset_statuses(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.AssetStatus).filter(models.AssetStatus.IsDeleted == 0).offset(skip).limit(limit).all()


def create_asset_status(db: Session, status_in: AssetStatusCreate):
    status = models.AssetStatus(**status_in.model_dump())
    db.add(status)
    db.commit()
    db.refresh(status)
    return status


def update_asset_status(db: Session, status_id: int, status_in: AssetStatusUpdate):
    status = get_asset_status(db, status_id)
    if not status:
        return None
    for field, value in status_in.model_dump(exclude_unset=True).items():
        setattr(status, field, value)
    db.commit()
    db.refresh(status)
    return status


def delete_asset_status(db: Session, status_id: int):
    status = get_asset_status(db, status_id)
    if not status:
        return None
    
    # Check if status is in use
    asset_count = db.query(models.Asset).filter(models.Asset.StatusId == status_id).count()
    if asset_count > 0:
        return {"error": f"Cannot delete status. It is assigned to {asset_count} asset(s)."}
    
    # Soft delete
    status.IsDeleted = 1
    status.DeletedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(status)
    return status


def get_asset(db: Session, asset_id: int):
    return db.query(models.Asset).filter(models.Asset.AssetId == asset_id, models.Asset.IsDeleted == 0).first()


def get_assets(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Asset).filter(models.Asset.IsDeleted == 0).offset(skip).limit(limit).all()


def create_asset(db: Session, asset_in: AssetCreate):
    asset = models.Asset(**asset_in.model_dump())
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def update_asset(db: Session, asset_id: int, asset_in: AssetUpdate):
    asset = get_asset(db, asset_id)
    if not asset:
        return None
    for field, value in asset_in.model_dump(exclude_unset=True).items():
        setattr(asset, field, value)
    db.commit()
    db.refresh(asset)
    return asset


def delete_asset(db: Session, asset_id: int):
    asset = get_asset(db, asset_id)
    if not asset:
        return None
    asset.IsDeleted = 1
    asset.DeletedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(asset)
    return asset


def get_asset_component(db: Session, component_id: int):
    return db.query(models.AssetComponent).filter(models.AssetComponent.ComponentId == component_id, models.AssetComponent.IsDeleted == 0).first()


def get_asset_components(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.AssetComponent).filter(models.AssetComponent.IsDeleted == 0).offset(skip).limit(limit).all()


def create_asset_component(db: Session, component_in: AssetComponentCreate):
    component = models.AssetComponent(**component_in.model_dump())
    db.add(component)
    db.commit()
    db.refresh(component)
    return component


def update_asset_component(db: Session, component_id: int, component_in: AssetComponentUpdate):
    component = get_asset_component(db, component_id)
    if not component:
        return None
    for field, value in component_in.model_dump(exclude_unset=True).items():
        setattr(component, field, value)
    db.commit()
    db.refresh(component)
    return component


def delete_asset_component(db: Session, component_id: int):
    component = get_asset_component(db, component_id)
    if not component:
        return None
    component.IsDeleted = 1
    component.DeletedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(component)
    return component


def get_asset_assignment(db: Session, assignment_id: int):
    return db.query(models.AssetAssignment).filter(models.AssetAssignment.AssignmentId == assignment_id).first()


def get_asset_assignments(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.AssetAssignment).offset(skip).limit(limit).all()


def assign_asset(db: Session, assignment_in: AssetAssignmentCreate):
    asset = get_asset(db, assignment_in.AssetId)
    if not asset:
        return None

    if asset.IsAvailable not in (None, 1):
        return None

    status_name = ''
    if asset.StatusId:
        status = get_asset_status(db, asset.StatusId)
        status_name = (status.StatusName or '').strip().lower() if status else ''
    if 'sold' in status_name:
        return None

    assignment = models.AssetAssignment(**assignment_in.model_dump())
    db.add(assignment)
    db.commit()
    db.refresh(assignment)

    if asset:
        asset.CurrentEmployeeId = assignment_in.EmployeeId
        asset.IsAvailable = 0
        db.commit()
        db.refresh(asset)

    history = models.AssetHistory(
        AssetId=assignment_in.AssetId,
        EmployeeId=assignment_in.EmployeeId,
        Action="assigned",
        Notes=assignment_in.Remarks,
        AssetCode=asset.AssetCode if asset else None,
        AssetName=asset.AssetName if asset else None,
        EmployeeName=db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == assignment_in.EmployeeId).first().FullName if db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == assignment_in.EmployeeId).first() else None,
    )
    db.add(history)
    db.commit()
    db.refresh(history)

    return assignment


def return_asset(db: Session, assignment_id: int, return_in: AssetAssignmentReturn):
    assignment = get_asset_assignment(db, assignment_id)
    if not assignment:
        return None
    assignment.ReturnedDate = return_in.ReturnedDate or datetime.datetime.utcnow().date()
    assignment.Remarks = return_in.Remarks
    assignment.IsReturned = 1
    db.commit()
    db.refresh(assignment)

    asset = get_asset(db, assignment.AssetId)
    if asset:
        asset.IsAvailable = 1
        asset.CurrentEmployeeId = None
        db.commit()
        db.refresh(asset)

    history = models.AssetHistory(
        AssetId=assignment.AssetId,
        EmployeeId=assignment.EmployeeId,
        Action="returned",
        Notes=return_in.Remarks,
        AssetCode=asset.AssetCode if asset else None,
        AssetName=asset.AssetName if asset else None,
        EmployeeName=db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == assignment.EmployeeId).first().FullName if db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == assignment.EmployeeId).first() else None,
    )
    db.add(history)
    db.commit()
    db.refresh(history)

    return assignment


def get_asset_history(db: Session, history_id: int):
    return db.query(models.AssetHistory).filter(models.AssetHistory.HistoryId == history_id).first()


def get_asset_histories(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.AssetHistory).offset(skip).limit(limit).all()


def create_asset_history(db: Session, history_in: AssetHistoryCreate):
    payload = history_in.model_dump()
    if not payload.get("AssetCode") and not payload.get("AssetName"):
        asset = get_asset(db, history_in.AssetId)
        if asset:
            payload["AssetCode"] = asset.AssetCode
            payload["AssetName"] = asset.AssetName
    if not payload.get("EmployeeName") and history_in.EmployeeId:
        employee = db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == history_in.EmployeeId).first()
        if employee:
            payload["EmployeeName"] = employee.FullName
    history = models.AssetHistory(**payload)
    db.add(history)
    db.commit()
    db.refresh(history)
    return history


# DetailedCategory CRUD
def get_detailed_category(db: Session, category_id: int):
    return db.query(models.DetailedCategory).filter(models.DetailedCategory.DetailedCategoryId == category_id, models.DetailedCategory.IsDeleted == 0).first()


def get_detailed_categories(db: Session, skip: int = 0, limit: int = 100):
    # Return hierarchical categories (top-level parents with children relationship)
    return db.query(models.DetailedCategory).filter(models.DetailedCategory.IsDeleted == 0, models.DetailedCategory.ParentId == None).offset(skip).limit(limit).all()


def create_detailed_category(db: Session, category_in: DetailedCategoryCreate):
    category = models.DetailedCategory(**category_in.model_dump())
    db.add(category)
    db.commit()
    db.refresh(category)
    return category


def update_detailed_category(db: Session, category_id: int, category_in: DetailedCategoryUpdate):
    category = db.query(models.DetailedCategory).filter(models.DetailedCategory.DetailedCategoryId == category_id).first()
    if not category:
        return None
    for field, value in category_in.model_dump(exclude_unset=True).items():
        setattr(category, field, value)
    db.commit()
    db.refresh(category)
    return category


def _get_detailed_category_subtree_ids(db: Session, category_id: int):
    category_ids = [category_id]
    child_ids = (
        db.query(models.DetailedCategory.DetailedCategoryId)
        .filter(models.DetailedCategory.ParentId == category_id, models.DetailedCategory.IsDeleted == 0)
        .all()
    )
    for (child_id,) in child_ids:
        category_ids.extend(_get_detailed_category_subtree_ids(db, child_id))
    return category_ids


def delete_detailed_category(db: Session, category_id: int):
    category = db.query(models.DetailedCategory).filter(models.DetailedCategory.DetailedCategoryId == category_id).first()
    if not category:
        return None

    category_ids = _get_detailed_category_subtree_ids(db, category_id)
    asset_count = db.query(models.DetailedAsset).filter(models.DetailedAsset.DetailedCategoryId.in_(category_ids)).count()
    if asset_count > 0:
        return {"error": f"Cannot delete detailed category. It is assigned to {asset_count} asset(s)."}

    category.IsDeleted = 1
    category.DeletedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(category)
    return category


def set_detailed_category_visibility(db: Session, category_id: int, is_hidden: bool, cascade: bool = True):
    """Hide or show a category. Hiding keeps the category and its assets, but the
    asset listings and pickers filter them out. Cascades to subcategories by default
    so hiding a parent does not leave visible orphans underneath it."""
    category = (
        db.query(models.DetailedCategory)
        .filter(models.DetailedCategory.DetailedCategoryId == category_id, models.DetailedCategory.IsDeleted == 0)
        .first()
    )
    if not category:
        return None

    target_ids = _get_detailed_category_subtree_ids(db, category_id) if cascade else [category_id]
    flag = 1 if is_hidden else 0
    for target_id in target_ids:
        target = (
            db.query(models.DetailedCategory)
            .filter(models.DetailedCategory.DetailedCategoryId == target_id)
            .first()
        )
        if target:
            target.IsHidden = flag

    db.commit()
    db.refresh(category)
    return category


def get_hidden_detailed_category_ids(db: Session) -> list[int]:
    rows = (
        db.query(models.DetailedCategory.DetailedCategoryId)
        .filter(models.DetailedCategory.IsDeleted == 0, models.DetailedCategory.IsHidden == 1)
        .all()
    )
    return [row[0] for row in rows]


def get_deleted_detailed_categories(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.DetailedCategory).filter(models.DetailedCategory.IsDeleted == 1).offset(skip).limit(limit).all()


def restore_detailed_category(db: Session, category_id: int):
    category = db.query(models.DetailedCategory).filter(models.DetailedCategory.DetailedCategoryId == category_id).first()
    if not category:
        return None
    category.IsDeleted = 0
    category.DeletedAt = None
    db.commit()
    db.refresh(category)
    return category


# DetailedAsset CRUD
def get_detailed_asset(db: Session, asset_id: int):
    return db.query(models.DetailedAsset).filter(models.DetailedAsset.DetailedAssetId == asset_id, models.DetailedAsset.IsDeleted == 0).first()


def get_detailed_assets(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.DetailedAsset).filter(models.DetailedAsset.IsDeleted == 0).offset(skip).limit(limit).all()


def _detailed_asset_unique_conflict(db: Session, data: dict, exclude_asset_id: int | None = None) -> str | None:
    """Message describing which unique value is already taken, or None when free.

    Asset tag and serial number are unique across every asset, deleted ones included, so
    the message says as much: a clash with a soft-deleted asset is otherwise baffling,
    since that asset is not in any listing. Checked here rather than left to the database
    so the caller gets a 409 with an explanation instead of a 500.
    """
    for field, label in (("AssetTag", "Asset tag"), ("SerialNo", "Serial number")):
        value = data.get(field)
        if value is None or not str(value).strip():
            continue
        query = db.query(models.DetailedAsset).filter(
            func.lower(getattr(models.DetailedAsset, field)) == str(value).strip().lower()
        )
        if exclude_asset_id is not None:
            query = query.filter(models.DetailedAsset.DetailedAssetId != exclude_asset_id)
        clash = query.first()
        if clash is None:
            continue
        if int(clash.IsDeleted or 0) == 1:
            return (
                f"{label} '{value}' is already held by a deleted asset, which is why it does not show in the "
                f"listing. Restore that asset from the Restore screen, or use a different {label.lower()}."
            )
        return f"{label} '{value}' is already used by asset {clash.AssetTag or clash.Name}."
    return None


def create_detailed_asset(db: Session, asset_in: DetailedAssetCreate):
    data = asset_in.model_dump()
    conflict = _detailed_asset_unique_conflict(db, data)
    if conflict:
        return {"error": conflict}
    asset = models.DetailedAsset(**data)
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def import_detailed_assets(db: Session, file: UploadFile):
    import csv
    from io import StringIO
    from datetime import datetime, date

    def parse_date(value):
        """Both dates are optional, so a blank cell reads as 'no date'. A real date cell
        from a spreadsheet arrives as a datetime, and one typed as text often carries a
        time component - accept all three rather than failing the row."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%m-%Y",
            "%d-%b-%Y",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
        ]:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            pass
        raise ValueError(f"Invalid date format: {value}")

    def parse_cost(value):
        """Tolerates a numeric cell, a thousands separator and a currency symbol."""
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text:
            return None
        cleaned = re.sub(r"[^0-9.\-]", "", text)
        try:
            return float(cleaned)
        except ValueError:
            raise ValueError(f"Invalid purchase cost: {value}")

    def detect_delimiter(text: str) -> str:
        """Spreadsheet exports use ';' or a tab in some locales. Match on whatever the
        header row uses, so this reader and the import screen's preview agree."""
        lines = text.splitlines()
        unquoted = re.sub(r'"[^"]*"', "", lines[0] if lines else "")
        counts = {delimiter: unquoted.count(delimiter) for delimiter in (",", ";", "\t")}
        best = max(counts, key=lambda delimiter: counts[delimiter])
        return best if counts[best] else ","

    def normalize_header(value) -> str:
        """Headers are matched on letters and digits only, so 'Make / Model',
        'make_model' and 'MakeModel' all resolve to the same field."""
        return re.sub(r"[^a-z0-9]", "", str(value if value is not None else "").strip().lower())

    # Accepts both the internal field names and the friendlier column titles used by
    # the downloadable template (ParentCategory / BrandName / Model).
    field_map = {
        "assettag": "AssetTag",
        "name": "Name",
        "brandname": "Name",
        "detailedcategoryid": "DetailedCategoryId",
        "detailedcategory": "DetailedCategoryId",
        "parentcategory": "DetailedCategoryId",
        "category": "DetailedCategoryId",
        "subcategory": "SubCategory",
        "makemodel": "MakeModel",
        "model": "MakeModel",
        "serialno": "SerialNo",
        "serialnumber": "SerialNo",
        "specifications": "Specifications",
        "specification": "Specifications",
        "status": "Status",
        "purchasecost": "PurchaseCost",
        "purchasedate": "PurchaseDate",
        "warrantyend": "WarrantyEnd",
        "customvalues": "CustomValues"
    }

    categories = {str(cat.DetailedCategoryId): cat for cat in db.query(models.DetailedCategory).filter(models.DetailedCategory.IsDeleted == 0).all()}
    categories_by_name = {cat.Name.strip().lower(): cat for cat in categories.values() if cat.Name}
    # A ParentCategory column names a top-level category, so resolve against those
    # first - a subcategory sharing the name must not win.
    parent_categories_by_name = {
        cat.Name.strip().lower(): cat for cat in categories.values() if cat.Name and cat.ParentId is None
    }

    # Asset tag and serial number are unique across every asset, deleted ones included -
    # a soft-deleted asset keeps its tag, so its clash has to be explained rather than
    # reported as a duplicate the importer cannot see.
    existing_tags: dict[str, models.DetailedAsset] = {}
    existing_serials: dict[str, models.DetailedAsset] = {}
    for asset in db.query(models.DetailedAsset).all():
        tag = str(asset.AssetTag or "").strip().lower()
        if tag:
            existing_tags[tag] = asset
        serial = str(asset.SerialNo or "").strip().lower()
        if serial:
            existing_serials[serial] = asset

    def duplicate_message(field: str, value, clash: models.DetailedAsset) -> str:
        if int(getattr(clash, "IsDeleted", 0) or 0) == 1:
            return (
                f"{field} '{value}' is already held by a deleted asset, which is why it does not show in the "
                f"listing. Restore that asset from the Restore screen, or give this row a different {field}."
            )
        return f"Duplicate {field} already exists: {value}"

    file_extension = file.filename.rsplit(".", 1)[-1].lower() if file.filename and "." in file.filename else ""
    rows = []
    # Values past the last header land under this key, which means the row's values no
    # longer line up with their columns - reported per row below rather than silently
    # dropped, so a stray comma in an unquoted value cannot corrupt an import.
    overflow_key = "__extra_values__"
    if file_extension == "csv":
        text_data = file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text_data), delimiter=detect_delimiter(text_data), restkey=overflow_key)
        rows = list(reader)
    elif file_extension in {"xls", "xlsx"}:
        try:
            import openpyxl
        except ImportError:
            raise ValueError("XLS/XLSX import requires openpyxl. Please install it before using this feature.")

        workbook = openpyxl.load_workbook(file.file, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(values_only=True), None) if sheet is not None else None
        if not header_row:
            raise ValueError("The spreadsheet has no header row.")
        header = [str(cell).strip() if cell is not None else "" for cell in header_row]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({header[idx]: row[idx] for idx in range(len(header)) if idx < len(row)})
    else:
        raise ValueError("Unsupported file type. Please upload a CSV or XLSX file.")

    created = 0
    processed = 0
    errors: list[dict] = []
    seen_tags: set[str] = set()
    seen_serials: set[str] = set()

    for index, raw_row in enumerate(rows, start=1):
        # A blank line carries no record, so it must not be counted or reported.
        if all(value is None or str(value).strip() == "" for value in raw_row.values() if not isinstance(value, list)):
            continue
        processed += 1
        extra_values = [value for value in (raw_row.get(overflow_key) or []) if str(value or "").strip()]
        row = {normalize_header(key): value for key, value in raw_row.items() if key and key != overflow_key}
        payload = {}
        try:
            if extra_values:
                raise ValueError(
                    "This row has more values than the file has columns, so its values no longer line up "
                    "with their fields. Wrap any value containing a comma - a JSON CustomValues cell, for "
                    "example - in double quotes."
                )
            for raw_key, raw_value in row.items():
                target = field_map.get(raw_key)
                if not target:
                    continue
                if target in {"PurchaseDate", "WarrantyEnd"}:
                    payload[target] = parse_date(raw_value)
                elif target == "PurchaseCost":
                    payload[target] = parse_cost(raw_value)
                elif target == "DetailedCategoryId":
                    if raw_value is None or str(raw_value).strip() == "":
                        payload[target] = None
                    elif str(raw_value).strip().isdigit():
                        payload[target] = int(str(raw_value).strip())
                    else:
                        lookup = str(raw_value).strip().lower()
                        match = parent_categories_by_name.get(lookup) or categories_by_name.get(lookup)
                        if not match:
                            raise ValueError(f"Parent category not found: {raw_value}")
                        payload[target] = match.DetailedCategoryId
                else:
                    # A blank cell means "no value", not an empty string. Asset tag and
                    # serial number are unique columns, where repeated empty strings clash
                    # but repeated NULLs do not.
                    text_value = str(raw_value).strip() if raw_value is not None else ""
                    payload[target] = text_value or None

            if "Name" not in payload or not payload["Name"]:
                raise ValueError("Missing required field 'Name'.")

            if payload.get("DetailedCategoryId") is not None and str(payload["DetailedCategoryId"]) not in categories:
                raise ValueError(f"Detailed category not found: {payload.get('DetailedCategoryId')}")

            if payload.get("AssetTag"):
                normalized_tag = str(payload["AssetTag"]).strip().lower()
                clash = existing_tags.get(normalized_tag)
                if clash is not None:
                    raise ValueError(duplicate_message("AssetTag", payload["AssetTag"], clash))
                if normalized_tag in seen_tags:
                    raise ValueError(f"Duplicate AssetTag in import file: {payload['AssetTag']}")
                seen_tags.add(normalized_tag)

            if payload.get("SerialNo"):
                normalized_serial = str(payload["SerialNo"]).strip().lower()
                clash = existing_serials.get(normalized_serial)
                if clash is not None:
                    raise ValueError(duplicate_message("SerialNo", payload["SerialNo"], clash))
                if normalized_serial in seen_serials:
                    raise ValueError(f"Duplicate SerialNo in import file: {payload['SerialNo']}")
                seen_serials.add(normalized_serial)

            asset_in = DetailedAssetCreate(**payload)
            create_detailed_asset(db, asset_in)
            created += 1
        except Exception as exc:
            errors.append({"row": index, "error": str(exc)})

    return {"processed": processed, "created": created, "errors": errors}


def update_detailed_asset(db: Session, asset_id: int, asset_in: DetailedAssetUpdate):
    asset = db.query(models.DetailedAsset).filter(models.DetailedAsset.DetailedAssetId == asset_id).first()
    if not asset:
        return None
    old_status = (asset.Status or '').strip().lower()
    data = asset_in.model_dump(exclude_unset=True)

    # Reject a tag or serial number already taken by another asset, rather than letting
    # the unique index fail the request with a 500.
    conflict = _detailed_asset_unique_conflict(db, data, exclude_asset_id=asset_id)
    if conflict:
        return {"error": conflict}

    # Lock sold status once sold price is set: allow sold price edits but prevent status change away from sold.
    status_locked = old_status == 'sold' and asset.SoldPrice is not None
    if status_locked and 'Status' in data:
        requested_status = (data.get('Status') or '').strip().lower()
        if requested_status and requested_status != 'sold':
            data['Status'] = asset.Status

    for field, value in data.items():
        setattr(asset, field, value)

    db.commit()
    db.refresh(asset)

    new_status = (asset.Status or '').strip().lower()
    maintenance_states = {"damaged", "damage", "maintenance"}
    # If asset became damaged/maintenance, create a workorder automatically
    if new_status in maintenance_states and old_status not in maintenance_states:
        # Create a workorder record
        ref = _generate_next_workorder_reference(db)
        reported_by_id = data.get('ReportedByEmployeeId') if isinstance(data.get('ReportedByEmployeeId'), int) else None
        reported_by_name = data.get('ReportedByName') if isinstance(data.get('ReportedByName'), str) else None
        wo = models.WorkOrder(
            Reference=ref,
            DetailedAssetId=asset.DetailedAssetId,
            ReportedByEmployeeId=reported_by_id,
            ReportedByName=reported_by_name,
            Status='open',
            Notes=f'Auto-created workorder for status change to {asset.Status}'
        )
        db.add(wo)
        db.commit()
        db.refresh(wo)

        # add detailed history entry
        hist = models.DetailedAssetHistory(
            DetailedAssetId=asset.DetailedAssetId,
            Action='marked-damaged',
            Notes=f'WorkOrder {wo.Reference} created',
            AssetTag=asset.AssetTag,
            AssetName=asset.Name
        )
        db.add(hist)
        db.commit()
        db.refresh(hist)

    return asset


### Vendor & WorkOrder CRUD
def create_vendor(db: Session, vendor_in):
    v = models.Vendor(**vendor_in.model_dump())
    db.add(v)
    db.commit()
    db.refresh(v)
    return v


def get_vendors(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Vendor).offset(skip).limit(limit).all()


def get_vendor(db: Session, vendor_id: int):
    return db.query(models.Vendor).filter(models.Vendor.VendorId == vendor_id).first()


def update_vendor(db: Session, vendor_id: int, vendor_in):
    vendor = get_vendor(db, vendor_id)
    if not vendor:
        return None
    for field, value in vendor_in.model_dump(exclude_unset=True).items():
        setattr(vendor, field, value)
    db.commit()
    db.refresh(vendor)
    return vendor


def create_workorder(db: Session, wo_in):
    data = wo_in.model_dump()
    if not data.get('Reference'):
        data['Reference'] = _generate_next_workorder_reference(db)
    if data.get('VendorId'):
        vendor = get_vendor(db, data['VendorId'])
        if vendor:
            data['VendorName'] = vendor.Name
    wo = models.WorkOrder(**data)
    db.add(wo)
    db.commit()
    db.refresh(wo)

    if data.get('DetailedAssetId'):
        asset = get_detailed_asset(db, data['DetailedAssetId'])
        # Raising a work order takes the asset in for repair: it comes back from whoever
        # was holding it, and returns to Available once the job is signed off. The asset's
        # status is the maintenance one - it used to be copied from the work order's own
        # status, which put values like "open" on the asset.
        send_asset_to_maintenance(db, asset, wo)

    return wo


CLOSED_WORKORDER_STATUSES = ('closed', 'completed', 'done', 'repaired')


def send_asset_to_maintenance(db: Session, asset, workorder, status: str | None = None):
    """Take an asset out of service for repair.

    An asset being repaired is not in the hands of the employee it was issued to, so any
    open assignment is closed off in their name with the work order recorded as the
    reason. Once the repair is signed off the asset becomes Available again, for Admin or
    HR to reassign - it is not handed straight back.
    """
    if asset is None:
        return []

    reference = getattr(workorder, "Reference", None) or "maintenance"
    open_assignments = (
        db.query(models.DetailedAssetAssignment)
        .filter(models.DetailedAssetAssignment.DetailedAssetId == asset.DetailedAssetId)
        .filter(models.DetailedAssetAssignment.IsReturned == 0)
        .all()
    )

    for assignment in open_assignments:
        assignment.IsReturned = 1
        assignment.ReturnedDate = datetime.date.today()
        assignment.ReturnedBy = f"Maintenance ({reference})"
        note = f"Returned for repair under work order {reference}"
        assignment.Remarks = f"{assignment.Remarks} | {note}" if assignment.Remarks else note

    asset.Status = (status or "").strip() or MAINTENANCE_STATUS
    db.commit()
    db.refresh(asset)

    # Recorded against each holder, so the return shows up in their own history rather
    # than the asset quietly vanishing from their list.
    for assignment in open_assignments:
        create_detailed_history(
            db,
            DetailedAssetHistoryCreate(
                DetailedAssetId=asset.DetailedAssetId,
                EmployeeId=assignment.EmployeeId,
                Action="Returned for Maintenance",
                Notes=f"Collected for repair under work order {reference}",
                AssetTag=asset.AssetTag,
                AssetName=asset.Name,
            ),
        )
    return open_assignments


def _holder_collected_from(db: Session, workorder):
    """The assignment closed when this job took the asset in, if there was one.

    `send_asset_to_maintenance` stamps the work order reference into ReturnedBy, which is
    what identifies the employee the asset came from - so it can go back to them once the
    repair is signed off.
    """
    if workorder is None or not workorder.Reference:
        return None
    return (
        db.query(models.DetailedAssetAssignment)
        .filter(models.DetailedAssetAssignment.DetailedAssetId == workorder.DetailedAssetId)
        .filter(models.DetailedAssetAssignment.ReturnedBy == f"Maintenance ({workorder.Reference})")
        .order_by(models.DetailedAssetAssignment.AssignmentId.desc())
        .first()
    )


def return_asset_after_repair(db: Session, asset, workorder, resolver):
    """Give a repaired asset back to the employee it was collected from.

    Returns the employee it went back to, or None when there is nobody to return it to -
    it was not assigned when it went in, someone else has been given it since, or that
    employee is no longer active - in which case the caller leaves it Available.
    """
    collected = _holder_collected_from(db, workorder)
    if asset is None or collected is None:
        return None

    already_out = (
        db.query(models.DetailedAssetAssignment)
        .filter(models.DetailedAssetAssignment.DetailedAssetId == asset.DetailedAssetId)
        .filter(models.DetailedAssetAssignment.IsReturned == 0)
        .first()
    )
    if already_out:
        return None

    employee = get_employee_detail(db, collected.EmployeeId)
    if employee is None or int(employee.IsActive or 0) == 0:
        return None

    reference = workorder.Reference or "maintenance"
    assignment = models.DetailedAssetAssignment(
        DetailedAssetId=asset.DetailedAssetId,
        EmployeeId=employee.EmployeeId,
        AssignedDate=datetime.date.today(),
        AssignedBy=(resolver.FullName or resolver.UserId) if resolver is not None else "Maintenance",
        Remarks=f"Returned after repair under work order {reference}",
        IsReturned=0,
    )
    db.add(assignment)
    db.commit()
    db.refresh(assignment)

    create_detailed_history(
        db,
        DetailedAssetHistoryCreate(
            DetailedAssetId=asset.DetailedAssetId,
            EmployeeId=employee.EmployeeId,
            Action="Returned After Repair",
            Notes=f"Handed back to {employee.FullName or employee.UserId} after work order {reference}",
            AssetTag=asset.AssetTag,
            AssetName=asset.Name,
            EmployeeName=employee.FullName or employee.UserId,
        ),
    )
    return employee


def status_after_maintenance(db: Session, asset) -> str:
    """Where an asset belongs once maintenance finishes.

    Normally Available: the asset was taken back from its holder when it went in for
    repair, and Admin or HR decide who gets it next. An assignment that is somehow still
    open - one recorded before this flow existed, say - means the asset never left that
    employee, so it stays Assigned rather than looking free while they hold it.
    """
    if asset is None:
        return AVAILABLE_STATUS
    still_assigned = (
        db.query(models.DetailedAssetAssignment)
        .filter(models.DetailedAssetAssignment.DetailedAssetId == asset.DetailedAssetId)
        .filter(models.DetailedAssetAssignment.IsReturned == 0)
        .first()
    )
    return ASSIGNED_STATUS if still_assigned else AVAILABLE_STATUS


def update_workorder(db: Session, workorder_id: int, wo_in):
    wo = db.query(models.WorkOrder).filter(models.WorkOrder.WorkOrderId == workorder_id).first()
    if not wo:
        return None
    for field, value in wo_in.model_dump(exclude_unset=True).items():
        setattr(wo, field, value)
    if wo.VendorId:
        vendor = get_vendor(db, wo.VendorId)
        if vendor:
            wo.VendorName = vendor.Name
    if getattr(wo, 'Status', None) and wo.Status.lower() in CLOSED_WORKORDER_STATUSES:
        wo.CompletedAt = datetime.datetime.utcnow()
        asset = get_detailed_asset(db, wo.DetailedAssetId)
        if asset:
            asset.Status = status_after_maintenance(db, asset)
            db.commit()
            db.refresh(asset)
    db.commit()
    db.refresh(wo)
    return wo


def get_workorder(db: Session, workorder_id: int):
    return db.query(models.WorkOrder).filter(models.WorkOrder.WorkOrderId == workorder_id).first()


def get_workorders(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.WorkOrder).offset(skip).limit(limit).all()


def get_workorders_for_asset(db: Session, detailed_asset_id: int):
    return db.query(models.WorkOrder).filter(models.WorkOrder.DetailedAssetId == detailed_asset_id).all()


def delete_detailed_asset(db: Session, asset_id: int):
    asset = db.query(models.DetailedAsset).filter(models.DetailedAsset.DetailedAssetId == asset_id).first()
    if not asset:
        return None
    asset.IsDeleted = 1
    asset.UpdatedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(asset)
    return asset


def get_deleted_detailed_assets(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.DetailedAsset).filter(models.DetailedAsset.IsDeleted == 1).offset(skip).limit(limit).all()


def restore_detailed_asset(db: Session, asset_id: int):
    asset = db.query(models.DetailedAsset).filter(models.DetailedAsset.DetailedAssetId == asset_id).first()
    if not asset:
        return None
    asset.IsDeleted = 0
    asset.UpdatedAt = datetime.datetime.utcnow()
    db.commit()
    db.refresh(asset)
    return asset


# DetailedAsset Assignment / History
def get_detailed_assignment(db: Session, assignment_id: int):
    return db.query(models.DetailedAssetAssignment).filter(models.DetailedAssetAssignment.AssignmentId == assignment_id).first()


def get_detailed_assignments(db: Session, detailed_asset_id: int | None = None, skip: int = 0, limit: int = 100):
    q = db.query(models.DetailedAssetAssignment)
    if detailed_asset_id is not None:
        q = q.filter(models.DetailedAssetAssignment.DetailedAssetId == detailed_asset_id)
    return q.offset(skip).limit(limit).all()


def assign_detailed_asset(db: Session, assignment_in: DetailedAssetAssignmentCreate):
    # Prevent assigning assets that are marked damaged, under maintenance, or sold
    asset = get_detailed_asset(db, assignment_in.DetailedAssetId)
    if asset and is_asset_out_of_service(asset.Status):
        return None

    # Prevent duplicate active assignment records for the same asset.
    open_assignment = (
        db.query(models.DetailedAssetAssignment)
        .filter(
            models.DetailedAssetAssignment.DetailedAssetId == assignment_in.DetailedAssetId,
            (models.DetailedAssetAssignment.IsReturned == 0) | (models.DetailedAssetAssignment.IsReturned.is_(None))
        )
        .first()
    )
    if open_assignment:
        return None

    assignment = models.DetailedAssetAssignment(**assignment_in.model_dump())
    db.add(assignment)
    db.commit()
    db.refresh(assignment)

    asset = get_detailed_asset(db, assignment_in.DetailedAssetId)
    if asset:
        # An asset that is out with an employee is no longer available.
        asset.Status = ASSIGNED_STATUS
        db.commit()
        db.refresh(asset)

    employee = db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == assignment_in.EmployeeId).first()
    notes = str(assignment_in.Remarks or '').strip()
    if assignment_in.AssignedBy:
        notes = f"Assigned by {assignment_in.AssignedBy}" + (f" - {notes}" if notes else "")
    history = models.DetailedAssetHistory(
        DetailedAssetId=assignment_in.DetailedAssetId,
        EmployeeId=assignment_in.EmployeeId,
        Action="assigned",
        Notes=notes,
        AssetTag=asset.AssetTag if asset else None,
        AssetName=asset.Name if asset else None,
        EmployeeName=employee.FullName if employee else None,
    )
    db.add(history)
    db.commit()
    db.refresh(history)

    return assignment


def return_detailed_asset(db: Session, assignment_id: int, return_in: DetailedAssetAssignmentReturn):
    assignment = get_detailed_assignment(db, assignment_id)
    if not assignment:
        return None
    assignment.ReturnedDate = return_in.ReturnedDate or datetime.datetime.utcnow().date()
    assignment.ReturnedBy = return_in.ReturnedBy
    assignment.Remarks = return_in.Remarks
    assignment.IsReturned = 1
    db.commit()
    db.refresh(assignment)

    asset = get_detailed_asset(db, assignment.DetailedAssetId)
    if asset:
        old_status = (asset.Status or '').strip().lower()
        requested_status = (return_in.Status or '').strip()
        if requested_status:
            asset.Status = requested_status
        elif old_status in {'', ASSIGNED_STATUS.lower()}:
            # No condition supplied - the asset is back on the shelf.
            asset.Status = AVAILABLE_STATUS
        db.commit()
        db.refresh(asset)

        # If the asset was returned and its new status is damaged/maintenance, auto-create a workorder
        maintenance_states = {"damaged", "damage", "maintenance"}
        new_status = (asset.Status or '').strip().lower()
        if new_status in maintenance_states and old_status not in maintenance_states:
            ref = _generate_next_workorder_reference(db)
            reported_by_name = return_in.ReturnedBy if isinstance(return_in.ReturnedBy, str) else None
            wo = models.WorkOrder(
                Reference=ref,
                DetailedAssetId=asset.DetailedAssetId,
                ReportedByEmployeeId=None,
                ReportedByName=reported_by_name,
                Status='open',
                Notes=f'Auto-created workorder for status change to {asset.Status} (returned)'
            )
            db.add(wo)
            db.commit()
            db.refresh(wo)

            # add detailed history entry
            hist = models.DetailedAssetHistory(
                DetailedAssetId=asset.DetailedAssetId,
                Action='marked-damaged',
                Notes=f'WorkOrder {wo.Reference} created (returned)',
                AssetTag=asset.AssetTag,
                AssetName=asset.Name
            )
            db.add(hist)
            db.commit()
            db.refresh(hist)

    employee = db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == assignment.EmployeeId).first()
    notes = str(return_in.Remarks or '').strip()
    if return_in.ReturnedBy:
        notes = f"{notes} (Returned by {return_in.ReturnedBy})" if notes else f"Returned by {return_in.ReturnedBy}"
    history = models.DetailedAssetHistory(
        DetailedAssetId=assignment.DetailedAssetId,
        EmployeeId=assignment.EmployeeId,
        Action="returned",
        Notes=notes,
        AssetTag=asset.AssetTag if asset else None,
        AssetName=asset.Name if asset else None,
        EmployeeName=employee.FullName if employee else None,
    )
    db.add(history)
    db.commit()
    db.refresh(history)

    return assignment


def get_detailed_history(db: Session, history_id: int):
    return db.query(models.DetailedAssetHistory).filter(models.DetailedAssetHistory.HistoryId == history_id).first()


def get_detailed_histories(db: Session, detailed_asset_id: int | None = None, skip: int = 0, limit: int = 100):
    q = db.query(models.DetailedAssetHistory)
    if detailed_asset_id is not None:
        q = q.filter(models.DetailedAssetHistory.DetailedAssetId == detailed_asset_id)
    return q.offset(skip).limit(limit).all()


def create_detailed_history(db: Session, history_in: DetailedAssetHistoryCreate):
    payload = history_in.model_dump()
    if not payload.get("AssetTag") and not payload.get("AssetName"):
        asset = get_detailed_asset(db, history_in.DetailedAssetId)
        if asset:
            payload["AssetTag"] = asset.AssetTag
            payload["AssetName"] = asset.Name
    if not payload.get("EmployeeName") and history_in.EmployeeId:
        employee = db.query(models.EmployeeDetail).filter(models.EmployeeDetail.EmployeeId == history_in.EmployeeId).first()
        if employee:
            payload["EmployeeName"] = employee.FullName
    history = models.DetailedAssetHistory(**payload)
    db.add(history)
    db.commit()
    db.refresh(history)
    return history


### Employee portal
# Every helper below takes the employee id resolved from the caller's token, so an
# employee can only ever read or create their own records.
WARRANTY_EXPIRY_WINDOW_DAYS = 30
# Statuses the procurement screen uses to close a request off - anything else still needs
# an answer, and counts as open on the requester's dashboard.
SETTLED_REQUEST_STATUSES = {"completed", "received", "rejected", "cancelled", "delivered"}


def _my_assignment_query(db: Session, employee_id: int):
    """Assignments belonging to one employee, newest first, with their asset joined."""
    return (
        db.query(models.DetailedAssetAssignment, models.DetailedAsset)
        .join(models.DetailedAsset, models.DetailedAsset.DetailedAssetId == models.DetailedAssetAssignment.DetailedAssetId)
        .filter(models.DetailedAssetAssignment.EmployeeId == employee_id)
        .filter(models.DetailedAsset.IsDeleted == 0)
        .order_by(
            models.DetailedAssetAssignment.AssignedDate.desc().nullslast(),
            models.DetailedAssetAssignment.AssignmentId.desc(),
        )
    )


def get_my_assets(db: Session, employee_id: int, include_returned: bool = False):
    """The assets an employee holds. Returned ones are past holdings, so they are left
    out unless asked for."""
    query = _my_assignment_query(db, employee_id)
    if not include_returned:
        query = query.filter(models.DetailedAssetAssignment.IsReturned == 0)

    category_names = {
        category.DetailedCategoryId: category.Name
        for category in db.query(models.DetailedCategory).all()
    }
    rows = []
    for assignment, asset in query.all():
        rows.append(
            {
                "AssignmentId": assignment.AssignmentId,
                "DetailedAssetId": asset.DetailedAssetId,
                "AssetTag": asset.AssetTag,
                "Name": asset.Name,
                "DetailedCategoryId": asset.DetailedCategoryId,
                "CategoryName": category_names.get(asset.DetailedCategoryId),
                "SubCategory": asset.SubCategory,
                "MakeModel": asset.MakeModel,
                "SerialNo": asset.SerialNo,
                "Specifications": asset.Specifications,
                "Status": asset.Status,
                "PurchaseDate": asset.PurchaseDate,
                "WarrantyEnd": asset.WarrantyEnd,
                "CustomValues": asset.CustomValues,
                "AssignedDate": assignment.AssignedDate,
                "ReturnedDate": assignment.ReturnedDate,
                "AssignedBy": assignment.AssignedBy,
                "ReturnedBy": assignment.ReturnedBy,
                "Remarks": assignment.Remarks,
                "IsReturned": int(assignment.IsReturned or 0),
            }
        )
    return rows


def get_my_history(db: Session, employee_id: int, skip: int = 0, limit: int = 200):
    return (
        db.query(models.DetailedAssetHistory)
        .filter(models.DetailedAssetHistory.EmployeeId == employee_id)
        .order_by(models.DetailedAssetHistory.ActionDate.desc(), models.DetailedAssetHistory.HistoryId.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )


def get_my_procurements(db: Session, employee_id: int):
    return (
        db.query(models.ProcurementRequest)
        .filter(models.ProcurementRequest.RequestedByEmployeeId == employee_id)
        .order_by(models.ProcurementRequest.ProcurementId.desc())
        .all()
    )


def create_my_procurement(db: Session, employee: models.EmployeeDetail, request_in):
    """Raise an equipment request on behalf of the employee making the call. Status is
    fixed to Pending here - only procurement staff move a request along."""
    data = request_in.model_dump()
    item = (data.get("Item") or "").strip()
    if not item:
        raise ValueError("Describe the equipment you need.")
    quantity = int(data.get("Quantity") or 1)
    if quantity < 1:
        raise ValueError("Quantity must be at least 1.")

    payload = ProcurementRequestCreate(
        Item=item,
        Quantity=quantity,
        CategoryId=data.get("CategoryId"),
        SubCategoryId=data.get("SubCategoryId"),
        Status="Pending",
        RequestedByEmployeeId=employee.EmployeeId,
        RequestedByName=employee.FullName or employee.UserId,
        Justification=(data.get("Justification") or "").strip() or None,
    )
    return create_procurement(db, payload)


def get_my_issues(db: Session, employee_id: int):
    return (
        db.query(models.WorkOrder)
        .filter(models.WorkOrder.ReportedByEmployeeId == employee_id)
        .order_by(models.WorkOrder.WorkOrderId.desc())
        .all()
    )


def create_my_issue(db: Session, employee: models.EmployeeDetail, issue_in):
    """Report a fault on an asset the employee is currently holding. Reporting against
    someone else's asset is refused, since the portal only covers your own equipment."""
    notes = (issue_in.Notes or "").strip()
    if not notes:
        raise ValueError("Describe the problem so it can be acted on.")

    holds_asset = (
        db.query(models.DetailedAssetAssignment)
        .filter(models.DetailedAssetAssignment.EmployeeId == employee.EmployeeId)
        .filter(models.DetailedAssetAssignment.DetailedAssetId == issue_in.DetailedAssetId)
        .filter(models.DetailedAssetAssignment.IsReturned == 0)
        .first()
    )
    if not holds_asset:
        raise ValueError("You can only report a problem for an asset assigned to you.")

    workorder = models.WorkOrder(
        Reference=_generate_next_workorder_reference(db),
        DetailedAssetId=issue_in.DetailedAssetId,
        ReportedByEmployeeId=employee.EmployeeId,
        ReportedByName=employee.FullName or employee.UserId,
        Status="open",
        Notes=notes,
    )
    db.add(workorder)
    db.commit()
    db.refresh(workorder)

    asset = get_detailed_asset(db, issue_in.DetailedAssetId)
    create_detailed_history(
        db,
        DetailedAssetHistoryCreate(
            DetailedAssetId=issue_in.DetailedAssetId,
            EmployeeId=employee.EmployeeId,
            Action="Issue Reported",
            Notes=f"WorkOrder {workorder.Reference}: {notes}",
            AssetTag=asset.AssetTag if asset else None,
            AssetName=asset.Name if asset else None,
            EmployeeName=employee.FullName or employee.UserId,
        ),
    )

    # The asset stays with the employee for now. It is only collected for repair when
    # someone starts the job from Employee Submissions, which is what moves it into
    # Maintenance and closes the assignment.
    return workorder


def get_my_portal_summary(db: Session, employee: models.EmployeeDetail):
    assets = get_my_assets(db, employee.EmployeeId, include_returned=True)
    held = [asset for asset in assets if not asset["IsReturned"]]
    today = datetime.date.today()
    horizon = today + datetime.timedelta(days=WARRANTY_EXPIRY_WINDOW_DAYS)
    open_statuses = {"open", "in progress", "pending"}

    return {
        "EmployeeId": employee.EmployeeId,
        "FullName": employee.FullName,
        "Department": employee.Department,
        "Designation": employee.Designation,
        "AssignedAssetCount": len(held),
        "ReturnedAssetCount": len(assets) - len(held),
        "OpenRequestCount": len(
            [
                request
                for request in get_my_procurements(db, employee.EmployeeId)
                if (request.Status or "").strip().lower() not in SETTLED_REQUEST_STATUSES
            ]
        ),
        "OpenIssueCount": len(
            [
                issue
                for issue in get_my_issues(db, employee.EmployeeId)
                if (issue.Status or "").strip().lower() in open_statuses
            ]
        ),
        "ExpiringWarrantyCount": len(
            [asset for asset in held if asset["WarrantyEnd"] and today <= asset["WarrantyEnd"] <= horizon]
        ),
    }


def start_workorder(db: Session, workorder_id: int, actor: models.EmployeeDetail):
    """Begin work on a reported fault: the asset is collected for repair.

    This is the step behind "Start" on Employee Submissions. Reporting a fault only
    records it - the asset stays with the employee until someone actually takes it in,
    which is here: the assignment is closed in their name and the asset moves into
    Maintenance, where it cannot be issued to anyone until the repair is signed off.
    """
    workorder = db.query(models.WorkOrder).filter(models.WorkOrder.WorkOrderId == workorder_id).first()
    if not workorder:
        return None

    workorder.Status = "in progress"
    asset = get_detailed_asset(db, workorder.DetailedAssetId)
    send_asset_to_maintenance(db, asset, workorder)
    db.commit()
    db.refresh(workorder)

    create_detailed_history(
        db,
        DetailedAssetHistoryCreate(
            DetailedAssetId=workorder.DetailedAssetId,
            EmployeeId=workorder.ReportedByEmployeeId,
            Action="Maintenance Started",
            Notes=(
                f"WorkOrder {workorder.Reference} taken in for repair by "
                f"{actor.FullName or actor.UserId}"
            ),
            AssetTag=asset.AssetTag if asset else None,
            AssetName=asset.Name if asset else None,
        ),
    )
    return workorder


def resolve_workorder(db: Session, workorder_id: int, resolver: models.EmployeeDetail, payload):
    """Close a reported issue off: who resolved it, when, and what was done.

    The repaired asset goes back to the employee it was collected from, as Assigned. If
    there is nobody to return it to it becomes Available instead, and either way the
    outcome is written to the asset's history so the employee who reported the fault can
    see what happened to it.
    """
    workorder = db.query(models.WorkOrder).filter(models.WorkOrder.WorkOrderId == workorder_id).first()
    if not workorder:
        return None

    data = payload.model_dump(exclude_unset=True) if payload is not None else {}
    workorder.Status = (data.get("Status") or "repaired").strip() or "repaired"
    workorder.Resolution = (data.get("Resolution") or "").strip() or None
    if data.get("RepairCost") is not None:
        workorder.RepairCost = data["RepairCost"]
    if data.get("VendorId") is not None:
        workorder.VendorId = data["VendorId"]
        vendor = get_vendor(db, data["VendorId"])
        if vendor:
            workorder.VendorName = vendor.Name
    workorder.ResolvedByEmployeeId = resolver.EmployeeId
    workorder.ResolvedByName = resolver.FullName or resolver.UserId
    workorder.CompletedAt = datetime.datetime.utcnow()

    asset = get_detailed_asset(db, workorder.DetailedAssetId)
    returned_to = None
    if asset:
        # A repaired asset goes back to the employee it was collected from, as Assigned.
        # With nobody to return it to it becomes Available - unless the caller names a
        # condition (a repaired asset may come back as "Good", for instance).
        returned_to = return_asset_after_repair(db, asset, workorder, resolver)
        requested_status = (data.get("AssetStatus") or "").strip()
        asset.Status = requested_status or status_after_maintenance(db, asset)

    db.commit()
    db.refresh(workorder)

    create_detailed_history(
        db,
        DetailedAssetHistoryCreate(
            DetailedAssetId=workorder.DetailedAssetId,
            EmployeeId=workorder.ReportedByEmployeeId,
            Action="Issue Resolved",
            Notes=(
                f"WorkOrder {workorder.Reference} resolved by {workorder.ResolvedByName}"
                + (f": {workorder.Resolution}" if workorder.Resolution else "")
                + (
                    f". Returned to {returned_to.FullName or returned_to.UserId}"
                    if returned_to is not None
                    else ""
                )
            ),
            AssetTag=asset.AssetTag if asset else None,
            AssetName=asset.Name if asset else None,
        ),
    )
    return workorder
